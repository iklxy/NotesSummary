#!/usr/bin/env python3
"""
@Date: 2026-05-08
@Author: lixinyang

Excel 导出工具。

当前实现基于 openpyxl 生成 CA 表格导出所需的 .xlsx。
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
except Exception:  # pragma: no cover - optional runtime support
    CellRichText = None  # type: ignore[assignment]
    TextBlock = None  # type: ignore[assignment]
    InlineFont = None  # type: ignore[assignment]

from interview_detail_fields import INTERVIEW_DETAIL_FIELD_LABELS


XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _clean_text(value: Any) -> str:
    """
    归一化单元格文本。
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n")
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _to_boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        text = value.strip().lower()
        return text in {"1", "true", "yes", "y", "on", "highlight"}
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def _normalize_answer_runs(raw_runs: Any) -> List[Dict[str, Any]]:
    """
    归一化答案高亮片段。
    """
    if not isinstance(raw_runs, list):
        return []
    runs: List[Dict[str, Any]] = []
    for item in raw_runs:
        if isinstance(item, str):
            text = item.strip()
            if text:
                runs.append({"text": text, "highlight": False})
            continue
        if not isinstance(item, dict):
            continue
        text = str(item.get("text") or item.get("value") or item.get("answer") or "").strip()
        if not text:
            continue
        runs.append(
            {
                "text": text,
                "highlight": _to_boolish(item.get("highlight") if item.get("highlight") is not None else item.get("emphasis")),
            }
        )
    return runs


def _build_answer_rich_text(answer_text: str, raw_runs: Any) -> Any:
    """
    根据 answer_runs 构造 openpyxl 富文本对象。
    """
    runs = _normalize_answer_runs(raw_runs)
    if not runs or CellRichText is None or TextBlock is None or InlineFont is None:
        return None
    rich_text = CellRichText()
    for run in runs:
        text = str(run.get("text") or "")
        if not text:
            continue
        highlight = bool(run.get("highlight"))
        font_kwargs: Dict[str, Any] = {"color": "FF111827"}
        if highlight:
            font_kwargs.update({"b": True, "color": "FFDC2626"})
        rich_text.append(TextBlock(InlineFont(**font_kwargs), text))
    if len(rich_text) == 0:
        return None
    return rich_text


def _make_answer_cell_spec(answer_text: str, raw_runs: Any) -> Dict[str, Any]:
    """
    构造答案单元格，必要时附加富文本渲染对象。
    """
    cell_spec: Dict[str, Any] = {"value": answer_text, "style": 1}
    rich_text = _build_answer_rich_text(answer_text, raw_runs)
    if rich_text is not None:
        cell_spec["rich_text"] = rich_text
    return cell_spec


def _normalize_ca_cell(value: Any) -> Dict[str, Any]:
    """
    归一化 CA 单元格。
    """
    if isinstance(value, dict):
        evidence_raw = value.get("evidence") or value.get("sources") or value.get("quotes") or []
        if isinstance(evidence_raw, list):
            evidence = [str(item or "").strip() for item in evidence_raw if str(item or "").strip()]
        elif isinstance(evidence_raw, str):
            evidence = [evidence_raw.strip()] if evidence_raw.strip() else []
        else:
            evidence = []
        answer = value.get("value")
        if answer is None:
            answer = value.get("answer")
        if answer is None:
            answer = value.get("text")
        answer_runs = _normalize_answer_runs(value.get("answer_runs") or value.get("answerRuns"))
        if (answer is None or str(answer).strip() == "") and answer_runs:
            answer = "".join(str(item.get("text") or "") for item in answer_runs)
        numeric_value_raw = value.get("numeric_value")
        if numeric_value_raw is None:
            numeric_value_raw = value.get("numericValue")
        numeric_value = None
        if isinstance(numeric_value_raw, (int, float)):
            numeric_value = float(numeric_value_raw)
        elif isinstance(numeric_value_raw, str):
            text = numeric_value_raw.strip()
            if text:
                try:
                    numeric_value = float(text)
                except Exception:
                    numeric_value = None
        return {
            "value": _clean_text(answer).strip() or "/",
            "evidence": evidence[:3],
            "answer_runs": answer_runs,
            "numeric_value": numeric_value,
        }
    return {"value": _clean_text(value).strip() or "/", "evidence": [], "answer_runs": [], "numeric_value": None}


def _build_ca_sheet_rows_v1(ca_payload: Dict[str, Any]) -> List[List[Dict[str, Any]]]:
    """
    兼容旧版 CA JSON 的 Excel 行数据。
    """
    project_id = ca_payload.get("project_id")
    project_name = str(ca_payload.get("project_name") or "").strip()
    questionnaire_name = str(ca_payload.get("questionnaire_name") or "").strip()
    generated_at = str(ca_payload.get("generated_at") or "").strip()
    selected_interview_ids = ca_payload.get("selected_interview_ids") or []
    column_meta_fields = ca_payload.get("column_meta_fields") or []
    column_meta_field_labels = ca_payload.get("column_meta_field_labels") or {}
    if not isinstance(column_meta_field_labels, dict):
        column_meta_field_labels = {}
    interviews = ca_payload.get("interviews") or []
    columns = ca_payload.get("columns") or []
    cells = ca_payload.get("cells") or {}

    interview_columns: List[Dict[str, Any]] = []
    interview_id_set = {str(item) for item in selected_interview_ids if item is not None}
    for item in interviews:
        if not isinstance(item, dict):
            continue
        interview_id = str(item.get("interview_id") or "").strip()
        name = str(item.get("name") or f"访谈 {interview_id}").strip()
        meta = item.get("meta") or {}
        meta_lines = [f"访谈ID：{interview_id}", f"名称：{name}"]
        if isinstance(meta, dict):
            for key in column_meta_fields:
                value = meta.get(key)
                label = str(column_meta_field_labels.get(key) or INTERVIEW_DETAIL_FIELD_LABELS.get(key) or key)
                if value is None or str(value).strip() == "":
                    meta_lines.append(f"{label}：/")
                else:
                    meta_lines.append(f"{label}：{value}")
        if not interview_id_set or interview_id in interview_id_set:
            interview_columns.append(
                {
                    "interview_id": interview_id,
                    "name": name,
                    "meta": meta,
                }
            )

    rows: List[List[Dict[str, Any]]] = []
    title_text = f"CA 表格 - {project_name or project_id}"
    rows.append([{"value": title_text, "style": 3}] + [{"value": "", "style": 3}] * len(interview_columns))
    rows.append(
        [
            {"value": f"项目：{project_name or project_id}", "style": 1},
            {"value": f"生成时间：{generated_at or ''}", "style": 1},
        ]
        + [{"value": "", "style": 1}] * max(0, len(interview_columns) - 2)
    )
    rows.append(
        [
            {"value": f"选择访谈：{len(interview_columns)}", "style": 1},
            {
                "value": "列字段："
                + (
                    ", ".join(
                        str(column_meta_field_labels.get(item) or INTERVIEW_DETAIL_FIELD_LABELS.get(item) or item)
                        for item in column_meta_fields
                        )
                        if column_meta_fields
                        else "无"
                ),
                "style": 1,
            },
        ]
        + [{"value": "", "style": 1}] * max(0, len(interview_columns) - 2)
    )
    rows.append([])

    rows.append([{"value": "访谈细节", "style": 3}] + [{"value": "", "style": 3}] * len(interview_columns))

    detail_specs = [
        ("访谈ID", lambda interview: str(interview.get("interview_id") or "").strip() or "/"),
        ("访谈名称", lambda interview: str(interview.get("name") or "").strip() or "/"),
        (
            "访谈日期",
            lambda interview: str(interview.get("interview_date") or "").split("T")[0] if interview.get("interview_date") else "-",
        ),
    ]
    for key in column_meta_fields:
        label = str(column_meta_field_labels.get(key) or INTERVIEW_DETAIL_FIELD_LABELS.get(key) or key)

        def _extract_meta_value(interview: Dict[str, Any], meta_key: str = key) -> str:
            meta = interview.get("meta") or {}
            if not isinstance(meta, dict):
                return "/"
            value = meta.get(meta_key)
            return _clean_text(value).strip() or "/"

        detail_specs.append((label, _extract_meta_value))

    for detail_label, extractor in detail_specs:
        rows.append(
            [{"value": detail_label, "style": 2}]
            + [{"value": extractor(column), "style": 1} for column in interview_columns]
        )

    rows.append([])
    rows.append([{"value": "问题", "style": 3}] + [{"value": "", "style": 3}] * len(interview_columns))

    for item in columns:
        if not isinstance(item, dict):
            continue
        if item.get("hidden"):
            continue
        question_uid = str(item.get("question_uid") or item.get("column_id") or "").strip()
        question_order = item.get("order")
        question_text = str(item.get("display_text") or item.get("question_text") or "").strip()
        if not question_uid and not question_text:
            continue
        row_cells: List[Dict[str, Any]] = [
            {
                "value": f"{question_order}. {question_text}" if question_order is not None else question_text,
                "style": 1,
            }
        ]
        for interview in interview_columns:
            interview_id = str(interview.get("interview_id") or "").strip()
            cell_text = "/"
            if interview_id and question_uid and isinstance(cells, dict):
                row_cells_by_interview = cells.get(interview_id)
                if isinstance(row_cells_by_interview, dict):
                    cell_text = _normalize_ca_cell(row_cells_by_interview.get(question_uid))["value"]
            row_cells.append({"value": cell_text, "style": 1})
        rows.append(row_cells)

    return rows


def _build_ca_sheet_rows_v2(
    ca_payload: Dict[str, Any],
    include_evidence_columns: bool = True,
) -> List[List[Dict[str, Any]]]:
    """
    将新版 CA JSON 转换为 Excel 行数据。
    """
    project_id = ca_payload.get("project_id")
    project_name = str(ca_payload.get("project_name") or "").strip()
    questionnaire_name = str(ca_payload.get("questionnaire_name") or "").strip()
    generated_at = str(ca_payload.get("generated_at") or "").strip()
    selected_interview_ids = ca_payload.get("selected_interview_ids") or []
    column_meta_fields = ca_payload.get("column_meta_fields") or []
    column_meta_field_labels = ca_payload.get("column_meta_field_labels") or {}
    if not isinstance(column_meta_field_labels, dict):
        column_meta_field_labels = {}
    interviews = ca_payload.get("interviews") or []
    columns = ca_payload.get("columns") or []
    cells = ca_payload.get("cells") or {}
    diff_row = ca_payload.get("diff_row") or {}

    interview_columns: List[Dict[str, Any]] = []
    interview_id_set = {str(item) for item in selected_interview_ids if item is not None}
    for item in interviews:
        if not isinstance(item, dict):
            continue
        interview_id = str(item.get("interview_id") or "").strip()
        name = str(item.get("name") or f"访谈 {interview_id}").strip()
        meta = item.get("meta") or {}
        if not interview_id_set or interview_id in interview_id_set:
            interview_columns.append(
                {
                    "interview_id": interview_id,
                    "name": name,
                    "meta": meta,
                }
            )

    cell_step = 2 if include_evidence_columns else 1
    total_cols = 1 + len(interview_columns) * cell_step
    rows: List[List[Dict[str, Any]]] = []
    title_text = f"CA 表格 - {project_name or project_id}"
    rows.append([{"value": title_text, "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    rows.append(
        [
            {"value": f"项目：{project_name or project_id}", "style": 1},
            {"value": f"生成时间：{generated_at or ''}", "style": 1},
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 2)
    )
    rows.append(
        [
            {"value": f"选择访谈：{len(interview_columns)}", "style": 1},
            {
                "value": "列字段："
                + (
                    ", ".join(
                        str(column_meta_field_labels.get(item) or INTERVIEW_DETAIL_FIELD_LABELS.get(item) or item)
                        for item in column_meta_fields
                    )
                    if column_meta_fields
                    else "无"
                ),
                "style": 1,
            },
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 2)
    )
    rows.append([])

    rows.append([{"value": "访谈细节", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))

    detail_specs = [
        ("访谈ID", lambda interview: str(interview.get("interview_id") or "").strip() or "/"),
        ("访谈名称", lambda interview: str(interview.get("name") or "").strip() or "/"),
        (
            "访谈日期",
            lambda interview: str(interview.get("interview_date") or "").split("T")[0] if interview.get("interview_date") else "-",
        ),
    ]
    for key in column_meta_fields:
        label = str(column_meta_field_labels.get(key) or INTERVIEW_DETAIL_FIELD_LABELS.get(key) or key)

        def _extract_meta_value(interview: Dict[str, Any], meta_key: str = key) -> str:
            meta = interview.get("meta") or {}
            if not isinstance(meta, dict):
                return "/"
            value = meta.get(meta_key)
            return _clean_text(value).strip() or "/"

        detail_specs.append((label, _extract_meta_value))

    for detail_label, extractor in detail_specs:
        row_values: List[Dict[str, Any]] = [{"value": detail_label, "style": 2}]
        for column in interview_columns:
            row_values.append({"value": extractor(column), "style": 1})
            if include_evidence_columns:
                row_values.append({"value": "", "style": 1})
        rows.append(row_values)

    rows.append([])
    rows.append([{"value": "问题", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))

    for item in columns:
        if not isinstance(item, dict):
            continue
        if item.get("hidden"):
            continue
        question_uid = str(item.get("question_uid") or item.get("column_id") or "").strip()
        question_order = item.get("order")
        question_text = str(item.get("display_text") or item.get("question_text") or "").strip()
        if not question_uid and not question_text:
            continue
        row_cells: List[Dict[str, Any]] = [
            {
                "value": f"{question_order}. {question_text}" if question_order is not None else question_text,
                "style": 1,
            }
        ]
        for interview in interview_columns:
            interview_id = str(interview.get("interview_id") or "").strip()
            answer_text = "/"
            evidence_text = ""
            answer_runs = []
            if interview_id and question_uid and isinstance(cells, dict):
                row_cells_by_interview = cells.get(interview_id)
                if isinstance(row_cells_by_interview, dict):
                    payload = _normalize_ca_cell(row_cells_by_interview.get(question_uid))
                    answer_text = payload["value"]
                    evidence_text = "\n".join(payload.get("evidence") or [])
                    answer_runs = payload.get("answer_runs") or []
            row_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
            if include_evidence_columns:
                row_cells.append({"value": evidence_text, "style": 1})
        rows.append(row_cells)

    rows.append([{"value": "差异化内容", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    diff_cells: List[Dict[str, Any]] = [{"value": "问卷未提及但访谈中出现的内容", "style": 1}]
    for interview in interview_columns:
        interview_id = str(interview.get("interview_id") or "").strip()
        answer_text = "/"
        evidence_text = ""
        answer_runs = []
        if interview_id and isinstance(diff_row, dict):
            payload = _normalize_ca_cell(diff_row.get(interview_id))
            answer_text = payload["value"]
            evidence_text = "\n".join(payload.get("evidence") or [])
            answer_runs = payload.get("answer_runs") or []
        diff_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
        if include_evidence_columns:
            diff_cells.append({"value": evidence_text, "style": 1})
    rows.append(diff_cells)

    return rows


def _format_number(value: float) -> str:
    """
    格式化统计数值。
    """
    if not isinstance(value, (int, float)) or not value == value:
        return "/"
    rounded = round(float(value), 2)
    if float(rounded).is_integer():
        return str(int(round(rounded)))
    return f"{rounded:.2f}".rstrip("0").rstrip(".")


def _extract_grouped_rows_v3(ca_payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    将新版 CA payload 中的列按主题分组展开。
    """
    columns = ca_payload.get("columns") or []
    if not isinstance(columns, list):
        columns = []
    explicit_groups = ca_payload.get("groups") or []
    grouped_rows: List[Dict[str, Any]] = []

    if isinstance(explicit_groups, list) and explicit_groups:
        sorted_groups = [
            item for item in explicit_groups if isinstance(item, dict)
        ]
        sorted_groups.sort(key=lambda item: int(item.get("order") or 0))
        consumed: set[str] = set()
        for group_index, group in enumerate(sorted_groups, start=1):
            label = str(group.get("title") or f"主题分组 {group_index}").strip() or f"主题分组 {group_index}"
            summary = str(group.get("summary") or "").strip()
            row_uids = [
                str(uid or "").strip()
                for uid in (group.get("row_uids") or [])
                if str(uid or "").strip()
            ]
            rows_in_group: List[Dict[str, Any]] = []
            if row_uids:
                row_uid_set = set(row_uids)
                for column in columns:
                    if not isinstance(column, dict):
                        continue
                    question_uid = str(column.get("question_uid") or column.get("column_id") or "").strip()
                    if question_uid in row_uid_set:
                        rows_in_group.append(column)
                        consumed.add(question_uid)
            else:
                rows_in_group = [column for column in columns if isinstance(column, dict) and str(column.get("group") or "").strip() == label]
                for column in rows_in_group:
                    question_uid = str(column.get("question_uid") or column.get("column_id") or "").strip()
                    if question_uid:
                        consumed.add(question_uid)
            if rows_in_group:
                grouped_rows.append(
                    {
                        "group_label": label,
                        "group_summary": summary,
                        "rows": rows_in_group,
                    }
                )
        leftovers = [
            column
            for column in columns
            if isinstance(column, dict)
            and str(column.get("question_uid") or column.get("column_id") or "").strip() not in consumed
        ]
        if leftovers:
            grouped_rows.append(
                {
                    "group_label": str(leftovers[0].get("group") or "未分组").strip() or "未分组",
                    "group_summary": str(leftovers[0].get("group_summary") or "").strip(),
                    "rows": leftovers,
                }
            )
        return grouped_rows

    current_label = None
    current_rows: List[Dict[str, Any]] = []
    current_summary = ""
    for column in columns:
        if not isinstance(column, dict):
            continue
        label = str(column.get("group") or "未分组").strip() or "未分组"
        if current_label is None or current_label != label:
            if current_rows:
                grouped_rows.append(
                    {
                        "group_label": current_label or "未分组",
                        "group_summary": current_summary,
                        "rows": current_rows,
                    }
                )
            current_label = label
            current_rows = [column]
            current_summary = str(column.get("group_summary") or "").strip()
            continue
        current_rows.append(column)
    if current_rows:
        grouped_rows.append(
            {
                "group_label": current_label or "未分组",
                "group_summary": current_summary,
                "rows": current_rows,
            }
        )
    return grouped_rows


def _build_ca_sheet_rows_v3(
    ca_payload: Dict[str, Any],
    include_evidence_columns: bool = True,
) -> List[List[Dict[str, Any]]]:
    """
    将 Notes 驱动的 CA JSON 转换为 Excel 行数据。
    """
    project_id = ca_payload.get("project_id")
    project_name = str(ca_payload.get("project_name") or "").strip()
    generated_at = str(ca_payload.get("generated_at") or "").strip()
    selected_interview_ids = ca_payload.get("selected_interview_ids") or []
    column_meta_fields = ca_payload.get("column_meta_fields") or []
    column_meta_field_labels = ca_payload.get("column_meta_field_labels") or {}
    if not isinstance(column_meta_field_labels, dict):
        column_meta_field_labels = {}
    interviews = ca_payload.get("interviews") or []
    diff_row = ca_payload.get("diff_row") or {}
    grouped_rows = _extract_grouped_rows_v3(ca_payload)

    interview_columns: List[Dict[str, Any]] = []
    interview_id_set = {str(item) for item in selected_interview_ids if item is not None}
    for item in interviews:
        if not isinstance(item, dict):
            continue
        interview_id = str(item.get("interview_id") or "").strip()
        name = str(item.get("name") or f"访谈 {interview_id}").strip()
        meta = item.get("meta") or {}
        if not interview_id_set or interview_id in interview_id_set:
            interview_columns.append(
                {
                    "interview_id": interview_id,
                    "name": name,
                    "meta": meta,
                }
            )

    cell_step = 2 if include_evidence_columns else 1
    total_cols = 2 + len(interview_columns) * cell_step
    rows: List[List[Dict[str, Any]]] = []
    title_text = f"CA Capture Sheet - {project_name or project_id}"
    rows.append([{"value": title_text, "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    rows.append(
        [
            {"value": f"项目：{project_name or project_id}", "style": 1},
            {"value": f"生成时间：{generated_at or ''}", "style": 1},
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 2)
    )
    rows.append(
        [
            {"value": f"选择访谈：{len(interview_columns)}", "style": 1},
            {
                "value": "列字段："
                + (
                    ", ".join(
                        str(column_meta_field_labels.get(item) or INTERVIEW_DETAIL_FIELD_LABELS.get(item) or item)
                        for item in column_meta_fields
                    )
                    if column_meta_fields
                    else "无"
                ),
                "style": 1,
            },
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 2)
    )
    rows.append([])

    rows.append([{"value": "访谈细节", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))

    detail_specs = [
        ("访谈ID", lambda interview: str(interview.get("interview_id") or "").strip() or "/"),
        ("访谈名称", lambda interview: str(interview.get("name") or "").strip() or "/"),
        (
            "访谈日期",
            lambda interview: str(interview.get("interview_date") or "").split("T")[0] if interview.get("interview_date") else "-",
        ),
    ]
    for key in column_meta_fields:
        label = str(column_meta_field_labels.get(key) or INTERVIEW_DETAIL_FIELD_LABELS.get(key) or key)

        def _extract_meta_value(interview: Dict[str, Any], meta_key: str = key) -> str:
            meta = interview.get("meta") or {}
            if not isinstance(meta, dict):
                return "/"
            value = meta.get(meta_key)
            return _clean_text(value).strip() or "/"

        detail_specs.append((label, _extract_meta_value))

    for detail_label, extractor in detail_specs:
        row_values: List[Dict[str, Any]] = [{"value": detail_label, "style": 2}]
        for column in interview_columns:
            row_values.append({"value": extractor(column), "style": 1})
            if include_evidence_columns:
                row_values.append({"value": "", "style": 1})
        rows.append(row_values)

    rows.append([])
    rows.append([{"value": "问题", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))

    for group in grouped_rows:
        group_label = str(group.get("group_label") or "未分组").strip() or "未分组"
        group_summary = str(group.get("group_summary") or "").strip()
        question_rows = group.get("rows") or []
        rows.append(
            [
                {"value": f"主题分组：{group_label}", "style": 3},
                {"value": group_summary or f"{len(question_rows)} 行", "style": 1},
            ]
            + [{"value": "", "style": 3}] * max(0, total_cols - 2)
        )
        for item in question_rows:
            if not isinstance(item, dict):
                continue
            if item.get("hidden"):
                continue
            question_uid = str(item.get("question_uid") or item.get("column_id") or "").strip()
            question_order = item.get("order")
            question_text = str(item.get("display_text") or item.get("question_text") or "").strip()
            question_type = str(item.get("question_type") or "qualitative").strip().lower()
            if not question_uid and not question_text:
                continue
            row_cells: List[Dict[str, Any]] = [
                {
                    "value": f"{question_order}. {question_text}" if question_order is not None else question_text,
                    "style": 1,
                },
            ]
            numeric_values: List[float] = []
            valid_count = 0
        for interview in interview_columns:
            interview_id = str(interview.get("interview_id") or "").strip()
            answer_text = "/"
            evidence_text = ""
            answer_runs = []
            if interview_id and question_uid and isinstance(ca_payload.get("cells"), dict):
                row_cells_by_interview = ca_payload.get("cells").get(interview_id)
                if isinstance(row_cells_by_interview, dict):
                    payload = _normalize_ca_cell(row_cells_by_interview.get(question_uid))
                    answer_text = payload["value"]
                    evidence_text = "\n".join(payload.get("evidence") or [])
                    answer_runs = payload.get("answer_runs") or []
                    value_text = str(payload.get("value") or "").strip()
                    if value_text and value_text != "/":
                        valid_count += 1
                        numeric_value = payload.get("numeric_value")
                        if isinstance(numeric_value, (int, float)):
                            numeric_values.append(float(numeric_value))
                        else:
                            parsed = _normalize_ca_cell({"value": value_text}).get("numeric_value")
                            if isinstance(parsed, (int, float)):
                                numeric_values.append(float(parsed))
                row_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
                if include_evidence_columns:
                    row_cells.append({"value": evidence_text, "style": 1})
            if question_type == "quantitative" and numeric_values:
                mean_value = sum(numeric_values) / len(numeric_values)
                min_value = min(numeric_values)
                max_value = max(numeric_values)
                stats_value = f"有效 {valid_count} / 均值 {_format_number(mean_value)} / 范围 {_format_number(min_value)}-{_format_number(max_value)}"
            else:
                stats_value = f"有效 {valid_count}"
            row_cells.insert(1, {"value": f"{stats_value}｜{'定量' if question_type == 'quantitative' else '定性'}", "style": 1})
            rows.append(row_cells)

    rows.append([{"value": "差异化内容", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    diff_valid_count = 0
    diff_numeric_values: List[float] = []
    diff_cells: List[Dict[str, Any]] = [
        {"value": "问卷未提及但访谈中出现的内容", "style": 1},
    ]
    for interview in interview_columns:
        interview_id = str(interview.get("interview_id") or "").strip()
        answer_text = "/"
        evidence_text = ""
        answer_runs = []
        if interview_id and isinstance(diff_row, dict):
            payload = _normalize_ca_cell(diff_row.get(interview_id))
            answer_text = payload["value"]
            evidence_text = "\n".join(payload.get("evidence") or [])
            answer_runs = payload.get("answer_runs") or []
            if answer_text and answer_text != "/":
                diff_valid_count += 1
            numeric_value = payload.get("numeric_value")
            if isinstance(numeric_value, (int, float)):
                diff_numeric_values.append(float(numeric_value))
        diff_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
        if include_evidence_columns:
            diff_cells.append({"value": evidence_text, "style": 1})
    if diff_numeric_values:
        diff_stats = f"有效 {diff_valid_count} / 均值 {_format_number(sum(diff_numeric_values) / len(diff_numeric_values))} / 范围 {_format_number(min(diff_numeric_values))}-{_format_number(max(diff_numeric_values))}"
    else:
        diff_stats = f"有效 {diff_valid_count}"
    diff_cells.insert(1, {"value": diff_stats, "style": 1})
    rows.append(diff_cells)

    return rows


def _build_ca_sheet_rows_v4(
    ca_payload: Dict[str, Any],
    include_evidence_columns: bool = True,
) -> List[List[Dict[str, Any]]]:
    """
    将带行总结列的 Notes 驱动 CA JSON 转换为 Excel 行数据。
    """
    project_id = ca_payload.get("project_id")
    project_name = str(ca_payload.get("project_name") or "").strip()
    generated_at = str(ca_payload.get("generated_at") or "").strip()
    selected_interview_ids = ca_payload.get("selected_interview_ids") or []
    column_meta_fields = ca_payload.get("column_meta_fields") or []
    column_meta_field_labels = ca_payload.get("column_meta_field_labels") or {}
    if not isinstance(column_meta_field_labels, dict):
        column_meta_field_labels = {}
    interviews = ca_payload.get("interviews") or []
    diff_row = ca_payload.get("diff_row") or {}
    grouped_rows = _extract_grouped_rows_v3(ca_payload)

    interview_columns: List[Dict[str, Any]] = []
    interview_id_set = {str(item) for item in selected_interview_ids if item is not None}
    for item in interviews:
        if not isinstance(item, dict):
            continue
        interview_id = str(item.get("interview_id") or "").strip()
        name = str(item.get("name") or f"访谈 {interview_id}").strip()
        meta = item.get("meta") or {}
        if not interview_id_set or interview_id in interview_id_set:
            interview_columns.append(
                {
                    "interview_id": interview_id,
                    "name": name,
                    "meta": meta,
                }
            )

    cell_step = 2 if include_evidence_columns else 1
    total_cols = 4 + len(interview_columns) * cell_step
    rows: List[List[Dict[str, Any]]] = []
    title_text = f"CA Capture Sheet - {project_name or project_id}"
    rows.append([{"value": title_text, "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    rows.append(
        [
            {"value": f"项目：{project_name or project_id}", "style": 1},
            {"value": f"生成时间：{generated_at or ''}", "style": 1},
            {"value": "", "style": 1},
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 3)
    )
    rows.append(
        [
            {"value": f"选择访谈：{len(interview_columns)}", "style": 1},
            {
                "value": "列字段："
                + (
                    ", ".join(
                        str(column_meta_field_labels.get(item) or INTERVIEW_DETAIL_FIELD_LABELS.get(item) or item)
                        for item in column_meta_fields
                    )
                    if column_meta_fields
                    else "无"
                ),
                "style": 1,
            },
            {"value": "", "style": 1},
        ]
        + [{"value": "", "style": 1}] * max(0, total_cols - 3)
    )
    rows.append([])

    rows.append([{"value": "访谈细节", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))

    detail_specs = [
        ("访谈ID", lambda interview: str(interview.get("interview_id") or "").strip() or "/"),
        ("访谈名称", lambda interview: str(interview.get("name") or "").strip() or "/"),
        (
            "访谈日期",
            lambda interview: str(interview.get("interview_date") or "").split("T")[0] if interview.get("interview_date") else "-",
        ),
    ]
    for key in column_meta_fields:
        label = str(column_meta_field_labels.get(key) or INTERVIEW_DETAIL_FIELD_LABELS.get(key) or key)

        def _extract_meta_value(interview: Dict[str, Any], meta_key: str = key) -> str:
            meta = interview.get("meta") or {}
            if not isinstance(meta, dict):
                return "/"
            value = meta.get(meta_key)
            return _clean_text(value).strip() or "/"

        detail_specs.append((label, _extract_meta_value))

    for detail_label, extractor in detail_specs:
        row_values: List[Dict[str, Any]] = [
            {"value": "", "style": 1},
            {"value": detail_label, "style": 2},
            {"value": "", "style": 1},
            {"value": "", "style": 1},
        ]
        for column in interview_columns:
            row_values.append({"value": extractor(column), "style": 1})
            if include_evidence_columns:
                row_values.append({"value": "", "style": 1})
        rows.append(row_values)

    rows.append([])
    rows.append([{"value": "问题", "style": 3}] + [{"value": "", "style": 3}] * (total_cols - 1))
    rows.append(
        [
            {"value": "主题分组", "style": 3},
            {"value": "问题内容", "style": 3},
            {"value": "有效答案统计", "style": 3},
            {"value": "单行总结", "style": 3},
        ]
        + [{"value": "", "style": 3}] * max(0, total_cols - 4)
    )

    for group in grouped_rows:
        group_label = str(group.get("group_label") or "未分组").strip() or "未分组"
        group_summary = str(group.get("group_summary") or "").strip()
        question_rows = group.get("rows") or []
        visible_question_rows: List[Dict[str, Any]] = []
        for item in question_rows:
            if not isinstance(item, dict):
                continue
            if item.get("hidden"):
                continue
            question_uid = str(item.get("question_uid") or item.get("column_id") or "").strip()
            question_order = item.get("order")
            question_text = str(item.get("display_text") or item.get("question_text") or "").strip()
            summary_text = str(item.get("summary_text") or "/").strip() or "/"
            if not question_uid and not question_text:
                continue
            visible_question_rows.append(item)
        if not visible_question_rows:
            continue
        for index, item in enumerate(visible_question_rows):
            question_uid = str(item.get("question_uid") or item.get("column_id") or "").strip()
            question_order = item.get("order")
            question_text = str(item.get("display_text") or item.get("question_text") or "").strip()
            question_type = str(item.get("question_type") or "qualitative").strip().lower()
            summary_text = str(item.get("summary_text") or "/").strip() or "/"
            row_cells: List[Dict[str, Any]] = [
                {
                    "value": group_label if index == 0 else "",
                    "style": 2,
                    "merge_span": len(visible_question_rows) if index == 0 else 0,
                },
                {
                    "value": f"{question_order}. {question_text}" if question_order is not None else question_text,
                    "style": 1,
                },
                {"value": "", "style": 1},
                {"value": summary_text, "style": 1},
            ]
            numeric_values: List[float] = []
            valid_count = 0
            for interview in interview_columns:
                interview_id = str(interview.get("interview_id") or "").strip()
                answer_text = "/"
                evidence_text = ""
                answer_runs = []
                if interview_id and question_uid and isinstance(ca_payload.get("cells"), dict):
                    row_cells_by_interview = ca_payload.get("cells").get(interview_id)
                    if isinstance(row_cells_by_interview, dict):
                        payload = _normalize_ca_cell(row_cells_by_interview.get(question_uid))
                        answer_text = payload["value"]
                        evidence_text = "\n".join(payload.get("evidence") or [])
                        answer_runs = payload.get("answer_runs") or []
                        value_text = str(payload.get("value") or "").strip()
                        if value_text and value_text != "/":
                            valid_count += 1
                            numeric_value = payload.get("numeric_value")
                            if isinstance(numeric_value, (int, float)):
                                numeric_values.append(float(numeric_value))
                            else:
                                parsed = _normalize_ca_cell({"value": value_text}).get("numeric_value")
                                if isinstance(parsed, (int, float)):
                                    numeric_values.append(float(parsed))
                row_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
                if include_evidence_columns:
                    row_cells.append({"value": evidence_text, "style": 1})
            if question_type == "quantitative" and numeric_values:
                mean_value = sum(numeric_values) / len(numeric_values)
                min_value = min(numeric_values)
                max_value = max(numeric_values)
                stats_value = f"有效 {valid_count} / 均值 {_format_number(mean_value)} / 范围 {_format_number(min_value)}-{_format_number(max_value)}"
            else:
                stats_value = f"有效 {valid_count}"
            row_cells[2] = {"value": stats_value, "style": 1}
            rows.append(row_cells)

    diff_valid_count = 0
    diff_numeric_values: List[float] = []
    diff_cells: List[Dict[str, Any]] = [
        {"value": "差异化内容", "style": 2},
        {"value": "问卷未提及但访谈中出现的内容", "style": 1},
        {"value": "", "style": 1},
        {"value": "", "style": 1},
    ]
    for interview in interview_columns:
        interview_id = str(interview.get("interview_id") or "").strip()
        answer_text = "/"
        evidence_text = ""
        answer_runs = []
        if interview_id and isinstance(diff_row, dict):
            payload = _normalize_ca_cell(diff_row.get(interview_id))
            answer_text = payload["value"]
            evidence_text = "\n".join(payload.get("evidence") or [])
            answer_runs = payload.get("answer_runs") or []
            if answer_text and answer_text != "/":
                diff_valid_count += 1
            numeric_value = payload.get("numeric_value")
            if isinstance(numeric_value, (int, float)):
                diff_numeric_values.append(float(numeric_value))
        diff_cells.append(_make_answer_cell_spec(answer_text, answer_runs))
        if include_evidence_columns:
            diff_cells.append({"value": evidence_text, "style": 1})
    if diff_numeric_values:
        diff_stats = f"有效 {diff_valid_count} / 均值 {_format_number(sum(diff_numeric_values) / len(diff_numeric_values))} / 范围 {_format_number(min(diff_numeric_values))}-{_format_number(max(diff_numeric_values))}"
    else:
        diff_stats = f"有效 {diff_valid_count}"
    diff_cells[2] = {"value": diff_stats, "style": 1}
    rows.append(diff_cells)

    return rows


def build_ca_table_xlsx_bytes(ca_payload: Dict[str, Any], include_evidence_columns: bool = True) -> bytes:
    """
    将 CA JSON 生成可下载的 Excel 二进制。
    """
    schema_version = int(ca_payload.get("schema_version") or 1)
    has_row_summary = False
    has_notes_layout = bool(ca_payload.get("groups"))
    if not has_notes_layout and isinstance(ca_payload.get("columns"), list):
        has_notes_layout = any(
            isinstance(column, dict)
            and (
                column.get("group")
                or column.get("group_id")
                or column.get("group_order")
                or column.get("group_summary")
                or column.get("question_type")
            )
            for column in ca_payload.get("columns")
        )
    if isinstance(ca_payload.get("columns"), list):
        has_row_summary = any(
            isinstance(column, dict) and column.get("summary_text") is not None
            for column in ca_payload.get("columns")
        )
    if schema_version >= 4 or has_row_summary:
        rows = _build_ca_sheet_rows_v4(ca_payload, include_evidence_columns=include_evidence_columns)
    elif schema_version >= 3 or has_notes_layout:
        rows = _build_ca_sheet_rows_v3(ca_payload, include_evidence_columns=include_evidence_columns)
    elif schema_version >= 2 or ca_payload.get("diff_row") is not None:
        rows = _build_ca_sheet_rows_v2(ca_payload, include_evidence_columns=include_evidence_columns)
    else:
        rows = _build_ca_sheet_rows_v1(ca_payload)
    total_cols = max(2, max((len(row) for row in rows), default=2))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CA表格"
    sheet.sheet_view.showGridLines = False
    is_summary_layout = schema_version >= 4 or has_row_summary
    sheet.freeze_panes = "E6" if is_summary_layout else ("C6" if (schema_version >= 3 or has_notes_layout) else "B6")
    sheet.sheet_view.zoomScale = 90

    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_index in range(1, total_cols + 1):
        letter = get_column_letter(col_index)
        if col_index == 1:
            sheet.column_dimensions[letter].width = 18 if is_summary_layout else 42
        elif is_summary_layout:
            if col_index == 2:
                sheet.column_dimensions[letter].width = 30
            elif col_index == 3:
                sheet.column_dimensions[letter].width = 28
            elif col_index == 4:
                sheet.column_dimensions[letter].width = 36
            elif include_evidence_columns and (col_index - 5) % 2 == 0:
                sheet.column_dimensions[letter].width = 36
            elif include_evidence_columns:
                sheet.column_dimensions[letter].width = 48
            else:
                sheet.column_dimensions[letter].width = 42
        elif schema_version >= 3 or has_notes_layout:
            if col_index == 2:
                sheet.column_dimensions[letter].width = 28
            elif include_evidence_columns and (col_index - 3) % 2 == 0:
                sheet.column_dimensions[letter].width = 36
            elif include_evidence_columns:
                sheet.column_dimensions[letter].width = 48
            else:
                sheet.column_dimensions[letter].width = 42
        elif schema_version >= 2:
            if include_evidence_columns:
                sheet.column_dimensions[letter].width = 34 if (col_index - 2) % 2 == 0 else 46
            else:
                sheet.column_dimensions[letter].width = 42
        else:
            sheet.column_dimensions[letter].width = 42 if schema_version >= 2 else 40

    current_section = None
    for row_index, row in enumerate(rows, start=1):
        first_value = _clean_text(row[0].get("value")) if row else ""
        second_value = _clean_text(row[1].get("value")) if len(row) > 1 else ""
        if first_value == "访谈细节":
            current_section = "detail"
        elif first_value == "问题":
            current_section = "question"
        elif first_value == "差异化内容":
            current_section = "diff"
        for col_index in range(1, total_cols + 1):
            if col_index <= len(row):
                cell_spec = row[col_index - 1]
                value = cell_spec.get("rich_text") if cell_spec.get("rich_text") is not None else cell_spec.get("value")
                style = int(cell_spec.get("style") or 1)
            else:
                value = ""
                style = 1
            cell = sheet.cell(row=row_index, column=col_index)
            if value is None:
                cell.value = ""
            elif isinstance(value, str):
                cell.value = _clean_text(value)
            else:
                cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = border
            if style == 3:
                cell.font = Font(name="Calibri", size=12, bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            elif style == 2:
                cell.font = Font(name="Calibri", size=12, bold=True)
            else:
                cell.font = Font(name="Calibri", size=12)

        if row_index == 1:
            sheet.row_dimensions[row_index].height = 32
        elif first_value in {"主题分组", "问题内容", "有效答案统计", "单行总结"}:
            sheet.row_dimensions[row_index].height = 32
        elif first_value in {"访谈细节", "问题"}:
            sheet.row_dimensions[row_index].height = 50
        elif current_section == "question":
            sheet.row_dimensions[row_index].height = 120
        elif current_section == "diff":
            sheet.row_dimensions[row_index].height = 120
        else:
            max_lines = 1
            for cell_spec in row:
                cell_text = _clean_text(cell_spec.get("value"))
                max_lines = max(max_lines, cell_text.count("\n") + 1)
            sheet.row_dimensions[row_index].height = min(max(24, max_lines * 22), 120)

    if is_summary_layout:
        for row_index, row in enumerate(rows, start=1):
            first_value = _clean_text(row[0].get("value")) if row else ""
            second_value = _clean_text(row[1].get("value")) if len(row) > 1 else ""
            if first_value == "问题" and second_value == "问题内容":
                continue
            if first_value == "差异化内容":
                continue
            if first_value == "主题分组" and second_value == "问题内容":
                continue
            merge_span = 0
            if row and isinstance(row[0], dict):
                merge_span_raw = row[0].get("merge_span")
                if isinstance(merge_span_raw, int):
                    merge_span = merge_span_raw
                elif isinstance(merge_span_raw, str) and merge_span_raw.strip():
                    try:
                        merge_span = int(merge_span_raw.strip())
                    except Exception:
                        merge_span = 0
            if merge_span > 1:
                end_row = row_index + merge_span - 1
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=end_row, end_column=1)
                merged_cell = sheet.cell(row=row_index, column=1)
                merged_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    if total_cols > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
